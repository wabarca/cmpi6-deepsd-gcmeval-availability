#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
CMIP6 Availability Inventory Builder
====================================

Genera una matriz de disponibilidad de variables CMIP6 a partir del
índice ESGF-West (MetaGrid).

Para cada combinación de:

    source_id
    variant_label
    experiment_id


determina si existen los datasets correspondientes a las variables
seleccionadas.

Salida:

    cmip6_daily_inventory.csv
    cmip6_daily_inventory.xlsx

Hojas Excel:

    inventory
        Disponibilidad detallada por modelo-realización-experimento.

    summary
        Resumen por modelo-realización.

    selected
        Modelos y realizaciones que poseen todas las variables para
        todos los experimentos solicitados.

Autor: Will Abarca (wabarca@ambiente.gob.sv)
"""

import requests
import pandas as pd
from collections import defaultdict
import json
from urllib.parse import quote

# ---------------------------------------------------------------------
# Configuración general
# ---------------------------------------------------------------------

# Endpoint utilizado por MetaGrid para consultar el índice ESGF
BASE_URL = "https://metagrid.esgf-west.org/proxy/search"

# Variables atmosféricas de interés
VARIABLES = [
    "ua",  # Viento zonal
    "va",  # Viento meridional
    "ta",  # Temperatura del aire
    "hur",  # Humedad relativa
    "hus",  # Humedad específica
    "zg",  # Altura geopotencial
    "psl",  # Presión reducida al nivel del mar
]

# Experimentos CMIP6 considerados
EXPERIMENTS = [
    "historical",
    "ssp126",
    "ssp245",
    "ssp370",
    "ssp585",
]

# Frecuencia temporal requerida
TABLE_ID = "day"

# Tamaño de página utilizado para paginación
PAGE_SIZE = 1000


def build_metagrid_url(source_id, variant_label):
    """
    Construye una URL de MetaGrid para la combinación
    modelo + realización seleccionada.
    """

    active_facets = {
        "table_id": "day",
        "experiment_id": [
            "historical",
            "ssp126",
            "ssp245",
            "ssp370",
            "ssp585",
        ],
        "variable_id": [
            "ua",
            "va",
            "ta",
            "hus",
            "hur",
            "zg",
            "psl",
        ],
        "frequency": "day",
        "source_id": source_id,
        "variant_label": variant_label,
    }

    return (
        "https://metagrid.esgf-west.org/search?"
        f"project=CMIP6&activeFacets="
        f"{quote(json.dumps(active_facets, separators=(',', ':')))}"
    )


def first(value):
    """
    Devuelve el primer elemento de una lista ESGF.

    Muchos campos del índice ESGF se devuelven como listas:

        source_id = ["EC-Earth3"]

    aunque conceptualmente contienen un único valor.

    Parameters
    ----------
    value : object

    Returns
    -------
    object
    """
    if isinstance(value, list):
        return value[0]

    return value


def fetch_variable_experiment(variable, experiment):
    """
    Recupera todos los datasets correspondientes a una combinación
    variable + experimento.

    Se utiliza paginación mediante offset y limit.

    Parameters
    ----------
    variable : str
        Variable CMIP6.

    experiment : str
        Experimento CMIP6.

    Returns
    -------
    list
        Lista de documentos ESGF.
    """

    offset = 0
    docs_all = []

    while True:

        params = {
            "project": "CMIP6",
            "table_id": TABLE_ID,
            "experiment_id": experiment,
            "variable_id": variable,
            # Mantener únicamente la versión más reciente
            "latest": "true",
            # Excluir réplicas en otros nodos ESGF
            # "data_node": "esgf-node.ornl.gov",
            "replica": "false",
            "type": "Dataset",
            "format": "application/solr+json",
            "limit": PAGE_SIZE,
            "offset": offset,
        }

        r = requests.get(
            BASE_URL,
            params=params,
            timeout=300,
        )

        r.raise_for_status()

        data = r.json()

        docs = data["response"]["docs"]

        if not docs:
            break

        docs_all.extend(docs)

        offset += len(docs)

        print(f"{experiment:10s} " f"{variable:5s} " f"{offset:5d}")

        if len(docs) < PAGE_SIZE:
            break

    return docs_all, data["response"]["numFound"]


def build_inventory():
    """
    Construye el inventario de disponibilidad.

    La clave de agregación utilizada es:

        (
            source_id,
            variant_label,
            experiment_id
        )

    Para cada clave se almacena el conjunto de variables encontradas.

    Returns
    -------
    tuple
        (
            inventory_dataframe,
            statistics_dict
        )
    """

    inventory = defaultdict(set)

    # Estadísticas de control
    total_numfound = 0
    total_downloaded = 0

    for experiment in EXPERIMENTS:

        print()
        print("=" * 70)
        print(f"Procesando experimento: {experiment}")
        print("=" * 70)

        for variable in VARIABLES:

            docs, num_found = fetch_variable_experiment(variable, experiment)

            if num_found > 9999:
                raise RuntimeError(
                    f"{experiment}-{variable}: "
                    f"{num_found} resultados exceden "
                    "el límite ESGF (9999)"
                )

            total_numfound += num_found
            total_downloaded += len(docs)

            print(
                f"{experiment:10s} "
                f"{variable:5s} "
                f"numFound={num_found:5d} "
                f"downloaded={len(docs):5d}"
            )

            for doc in docs:
                source_id = first(doc.get("source_id"))

                variant_label = first(doc.get("variant_label"))

                experiment_id = first(doc.get("experiment_id"))

                variable_id = first(doc.get("variable_id"))

                key = (
                    source_id,
                    variant_label,
                    experiment_id,
                )

                inventory[key].add(variable_id)

    rows = []

    for key, vars_found in inventory.items():

        (
            source_id,
            variant_label,
            experiment_id,
        ) = key

        row = {
            "source_id": source_id,
            "variant_label": variant_label,
            "experiment_id": experiment_id,
        }

        for var in VARIABLES:
            row[var] = var in vars_found

        row["score"] = len(vars_found)

        row["complete"] = row["score"] == len(VARIABLES)

        rows.append(row)

    df = pd.DataFrame(rows)

    df = df.sort_values(
        [
            "source_id",
            "variant_label",
            "experiment_id",
        ]
    )

    stats = {
        "total_numfound": total_numfound,
        "total_downloaded": total_downloaded,
        "inventory_rows": len(df),
        "unique_models": df["source_id"].nunique(),
    }

    print()
    print("=" * 70)
    print("RESUMEN ESGF")
    print("=" * 70)
    print(f"Documentos reportados por ESGF : " f"{stats['total_numfound']:,}")
    print(f"Documentos descargados         : " f"{stats['total_downloaded']:,}")
    print(f"Filas inventario              : " f"{stats['inventory_rows']:,}")
    print(f"Modelos únicos                : " f"{stats['unique_models']:,}")
    print("=" * 70)
    print()

    return df, stats


def build_summary(df):
    """
    Genera un resumen por modelo-realización.
    """

    max_possible = len(VARIABLES) * len(EXPERIMENTS)

    summary = (
        df.groupby(["source_id", "variant_label"])
        .agg(
            complete_experiments=("complete", "sum"),
            total_variables=("score", "sum"),
        )
        .reset_index()
    )

    summary["availability_pct"] = 100.0 * summary["total_variables"] / max_possible

    summary["all_experiments_complete"] = summary["complete_experiments"] == len(
        EXPERIMENTS
    )

    summary = summary.sort_values(
        [
            "complete_experiments",
            "total_variables",
            "availability_pct",
        ],
        ascending=False,
    )

    return summary


def main():
    """
    Flujo principal del programa.
    """

    print()
    print("Construyendo inventario CMIP6...")
    print()

    df, stats = build_inventory()

    summary = build_summary(df)

    # -------------------------------------------------
    # Comparación con modelos disponibles en GCMEval
    # -------------------------------------------------

    gcmeval = pd.read_csv("gcmeval_models.csv")

    gcmeval["model_key"] = (
        gcmeval["source_id"].astype(str) + "|" + gcmeval["variant_label"].astype(str)
    )

    summary["model_key"] = (
        summary["source_id"].astype(str) + "|" + summary["variant_label"].astype(str)
    )

    gcmeval_set = set(gcmeval["model_key"])

    summary["gcmeval"] = summary["model_key"].isin(gcmeval_set)

    summary = summary.drop(columns=["model_key"])

    # -------------------------------------------------
    # Selección final
    # -------------------------------------------------

    selected = summary[summary["complete_experiments"] == len(EXPERIMENTS)].copy()

    selected["dataset_url"] = selected.apply(
        lambda row: build_metagrid_url(
            row["source_id"],
            row["variant_label"],
        ),
        axis=1,
    )

    selected["metagrid"] = "Abrir"

    csv_file = "cmip6_daily_inventory.csv"
    xlsx_file = "cmip6_daily_inventory.xlsx"

    df.to_csv(csv_file, index=False)

    with pd.ExcelWriter(xlsx_file, engine="openpyxl") as writer:

        # -------------------------------------------------
        # Exportar hojas
        # -------------------------------------------------

        df.to_excel(writer, sheet_name="inventory", index=False)

        summary.to_excel(writer, sheet_name="summary", index=False)

        selected.to_excel(writer, sheet_name="selected", index=False)

        from openpyxl.styles import PatternFill

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE",
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC7CE",
            end_color="FFC7CE",
        )

        # =================================================
        # INVENTORY
        # =================================================

        ws_inventory = writer.sheets["inventory"]

        variable_cols = {}
        complete_col = None

        for col_num, cell in enumerate(ws_inventory[1], start=1):

            if cell.value in VARIABLES:
                variable_cols[cell.value] = col_num

            elif cell.value == "complete":
                complete_col = col_num

        # Colorear variables
        for col_num in variable_cols.values():

            for row in range(2, ws_inventory.max_row + 1):

                cell = ws_inventory.cell(row=row, column=col_num)

                if cell.value is True:
                    cell.fill = green_fill

                elif cell.value is False:
                    cell.fill = red_fill

        # Colorear columna complete
        if complete_col:

            for row in range(2, ws_inventory.max_row + 1):

                cell = ws_inventory.cell(row=row, column=complete_col)

                if cell.value is True:
                    cell.fill = green_fill

                elif cell.value is False:
                    cell.fill = red_fill

        # =================================================
        # SELECTED
        # =================================================

        ws = writer.sheets["selected"]

        metagrid_col = None
        url_col = None
        gcmeval_col = None

        for col_num, cell in enumerate(ws[1], start=1):

            if cell.value == "metagrid":
                metagrid_col = col_num

            elif cell.value == "dataset_url":
                url_col = col_num

            elif cell.value == "gcmeval":
                gcmeval_col = col_num

        # Hipervínculos MetaGrid
        if metagrid_col and url_col:

            for row in range(2, ws.max_row + 1):

                link_cell = ws.cell(row=row, column=metagrid_col)

                url_cell = ws.cell(row=row, column=url_col)

                if url_cell.value:

                    link_cell.hyperlink = url_cell.value
                    link_cell.style = "Hyperlink"

            from openpyxl.utils import get_column_letter

            ws.column_dimensions[get_column_letter(url_col)].hidden = True

        # Colorear columna gcmeval
        if gcmeval_col:

            for row in range(2, ws.max_row + 1):

                cell = ws.cell(row=row, column=gcmeval_col)

                if cell.value is True:
                    cell.fill = green_fill

                elif cell.value is False:
                    cell.fill = red_fill

    print()
    print("=" * 70)
    print("FINALIZADO")
    print("=" * 70)
    print(f"CSV                     : {csv_file}")
    print(f"Excel                   : {xlsx_file}")
    print(f"Documentos ESGF         : " f"{stats['total_numfound']:,}")
    print(f"Documentos descargados  : " f"{stats['total_downloaded']:,}")
    print(f"Filas inventario        : " f"{stats['inventory_rows']:,}")
    print(f"Modelos únicos          : " f"{stats['unique_models']:,}")
    print(f"Seleccionados           : " f"{len(selected):,}")
    print(f"Disponibles en GCMEval  : " f"{summary['gcmeval'].sum():,}")
    print(f"Seleccionados + GCMEval : " f"{selected['gcmeval'].sum():,}")


if __name__ == "__main__":
    main()
